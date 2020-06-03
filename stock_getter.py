import logging
import openpyxl

from yahoo_fin import stock_info as si


def main():
    wb = openpyxl.load_workbook('stocks.xlsx')
    sheet = wb.active
    for i in range(2, 50):
        ticker = sheet.cell(column=1, row=i).value
        if not ticker:
            break
        price = si.get_live_price(str(ticker)).round(2)
        sheet.cell(column=6, row=i).value = price
        print(f'Setting {ticker} to price: {price}')
    wb.save('stocks.xlsx')

if __name__ == "__main__":
    main()